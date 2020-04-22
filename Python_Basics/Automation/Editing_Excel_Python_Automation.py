import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from os import walk


def get_Filenames(path):
    f = []
    #fetching file names from walk method imported from os
    for (dirpath, dirnames, filenames) in walk(path):
        f.extend(filenames)
        # f = [x for x in f if x.endswith(".xlsx")]
        f = [x for x in f if (".xls") in x]
    return f

def WorkBook_Editing(filename):
    workbook = xl.load_workbook(filename)
    sheet = workbook['Sheet1']
    cell = sheet.cell(1,1)
    #print(cell)
    #print(sheet.max_row)

    for i in range(2,sheet.max_row+1):
        sheet.cell(i,4).value =sheet.cell(i,3).value*1.5

    #Adding chart to the excel using openpyxl

    #restricting the values of the chart to the column 4
    values = Reference(sheet,min_row=2,max_row=4,min_col=4,max_col=4)
    chart = BarChart()

    #adding values to chart from reference object
    chart.add_data(values)

    #adding the chart to a particular cell
    sheet.add_chart(chart,'a6')

    #saving with same file instead of creating a new copy(give new name for new copy)
    workbook.save(filename)

mypath = 'C:\\Users\\yella\\Desktop\\Python_code\\Python_Basics\\Automation'
filenames = get_Filenames(mypath)
for file in filenames:
    WorkBook_Editing(file)